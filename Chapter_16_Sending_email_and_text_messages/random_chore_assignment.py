'''
Write a program that takes a list of people’s email addresses and a list
of chores that need to be done and randomly assigns chores to people.
Email each person their assigned chores.
'''

import os
import openpyxl
import random
import smtplib


def random_tasks_assignment(file_name, project_name, emails_sheet, tasks_sheet, tasks_per_email, assignments_sheet,
                            from_email, password, cwd=None):
    """
    Randomly assigns task to an email and sends email notification with such assignment.
    Also, it records emails-assigned tasks pairs to the input Excel file.
    Prerequisite: Excel file with 2 sheets - emails and tasks list.
    :param file_name: string, valid file name
    :param project_name: string
    :param emails_sheet: string, valid sheet name
    :param tasks_sheet: string, valid sheet name
    :param tasks_per_email: int
    :param assignments_sheet: string
    :param from_email: string, valid email
    :param password: string, valid password
    :param cwd: optional, absolute path
    :return:
    """

    # switch to the cwd path if exists
    if cwd:
        # if cwd is passed, check that path is valid
        if not os.path.isdir(cwd):
            print(f'Path "{cwd}" does not exists.\nExiting the program.')
            return
        os.chdir(cwd)
    else:
        # if cwd is not passes, use current directory
        os.chdir('.')

    # load workbook if exists
    if file_name not in os.listdir(cwd):
        print(f"File '{file_name}' does not exist in path {cwd}.\nExiting program.")
        return

    wb = openpyxl.load_workbook(file_name)

    # verify that data sheets exist
    if emails_sheet not in wb.sheetnames or tasks_sheet not in wb.sheetnames:
        print(f"File doesn't have sheet with name '{emails_sheet}' or '{tasks_sheet}'.\nExiting program.")

    # extract emails into variable
    sheet = wb[emails_sheet]
    emails = []
    for cell in sheet['A']:
        if cell.value:  # skip None values
            emails.append(cell.value)

    # extract tasks into variable
    sheet = wb[tasks_sheet]
    tasks = []
    for cell in sheet['A']:
        if cell.value:  # skip None values
            tasks.append(cell.value)

    # verify all tasks can be assigned
    if tasks_per_email*len(emails) != len(tasks):
        print(f"Not all tasks can be matched. Emails number - {len(emails)}, tasks per person - {tasks_per_email}, "
              f"tasks number = {len(tasks)}. Please enter correct values.\nExiting program.")
        return

    # assign tasks
    assigned_tasks = {}
    for email in emails:
        random_assignment = random.randint(0, len(tasks)-1)
        assigned_tasks.setdefault(email, tasks[random_assignment])
        del tasks[random_assignment]

    # record assigned tasks to excel
    if assignments_sheet in wb.sheetnames:
        del wb[assignments_sheet]

    wb.create_sheet(assignments_sheet)
    sheet = wb[assignments_sheet]
    for i, (email, task) in enumerate(assigned_tasks.items()):
        sheet[f"A{i+1}"] = email
        sheet[f"B{i+1}"] = task
    wb.save(file_name)
    print(f"Recorded email-assignments pairs into new sheet '{assignments_sheet}'")

    # connect to email server
    try:
        smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
        smtp_obj.ehlo()
        smtp_obj.starttls()
        smtp_obj.login(from_email, password)
    except Exception as Ex:
        print(f"Exception while connecting to mail server: {Ex}")
    else:

        # send emails
        for to_email, task in assigned_tasks.items():
            send_email = smtp_obj.sendmail(from_email, to_email,
                                           f"Subject: Task assignment for '{project_name}'\n\n"
                                           f"Hi there!\nYou've been randomly assigned next task - {task}.\n"
                                           "Thanks in advance for your involvement and effort!")
            if send_email == {}:
                print(f"Email to {to_email} was sent successfully")
            else:
                print(f"Failed to sent email to {to_email}")


active_directory = r"D:\Practice Python"
file = "random_chore_assignment_file.xlsx"
project = "StarNight Gala Event"
emails_sheet_list = "Emails"
tasks_sheet_list = "Tasks"
tasks_per_person = 1
result_sheet = "Assigned_tasks"
sender_email = "******@gmail.com"
email_password = "********"

random_tasks_assignment(file, project, emails_sheet_list, tasks_sheet_list, tasks_per_person,
                        result_sheet, sender_email, email_password)

