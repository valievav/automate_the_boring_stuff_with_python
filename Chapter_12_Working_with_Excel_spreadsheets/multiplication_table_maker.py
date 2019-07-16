'''
Create a program multiplicationTable.py that takes a number N from the command
line and creates an N×N multiplication table in an Excel spreadsheet.
Row 1 and column A should be used for labels and should be in bold.
Example of the program: py multiplicationTable.py 6
'''

import openpyxl, os, sys
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font


def multiplication_table(file_name, cwd=None):
    """
    Uses multiplication number N from the command line to build\n
    a multiplication table NxN in Excel spreadsheet.
    :param cwd: absolute path
    :param file_name: file name with .xlsx extension
    :return: .xlsx file
    """

    if cwd:
        os.chdir(working_folder)

    # get multiplication number from cmd
    try:
        multiplication_number = int(sys.argv[1])
    except (ValueError, IndexError):
        print("Provided incorrect multiplication number. Please make sure it's a digit.")
        sys.exit()
    else:

        # create Excel workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        bold_font = Font(bold=True)

        # fill in column A and row 1
        for number in range(1, multiplication_number + 1):
            sheet["A" + str(number + 1)] = number  # column A
            sheet[get_column_letter(number + 1) + "1"] = number  # row 1
            sheet["A" + str(number + 1)].font = sheet[get_column_letter(number + 1) + "1"].font = bold_font

        # fill in calc values
        for column_number in range(2, multiplication_number+2):
            for row_number in range(2, multiplication_number+2):
                col_A_cell = get_column_letter(1)+str(row_number)
                row_1_cell = get_column_letter(column_number) + str(1)
                sheet[get_column_letter(column_number) + str(row_number)] = f'={col_A_cell}*{row_1_cell}'

        wb.save(file_name)
        print(f"Created file with {multiplication_number}x{multiplication_number} multiplication table.\n"
              f"File name: {file_name}\n"
              f"File location: {os.getcwd()}\n")


working_folder = "D:\Chapter 12 - Excel files"
result_file_name = "multiplication_table.xlsx"

multiplication_table(result_file_name)

