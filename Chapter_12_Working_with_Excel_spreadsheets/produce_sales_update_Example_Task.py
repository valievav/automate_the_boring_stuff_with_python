import openpyxl, os
from openpyxl.utils.cell import get_column_letter


def update_produce_price(file, produce_column, price_column, produce_price, new_file_name, cwd=None):
    """
    Updates produce prices in the Excel spreadsheet to the correct values.\n
    :param file: xlsx file
    :param produce_column: int
    :param price_column:  int
    :param produce_price: dictionary where keys - produce names and values - new prices
    :param new_file_name: xlsx file
    :param cwd: absolute path for working folder (optional param)
    :return: generates new file with the updated data
    """

    # can work with files in different directory than source code
    if cwd:
        if os.path.isabs(cwd):
            os.chdir(cwd)
        else:
            print("Please enter absolute path for cwd parameter or exclude from function arguments.")
            return

    if not (file.endswith(".xlsx") and new_file_name.endswith(".xlsx")):
        print("Cannot process files with extension different from '.xlsx'.")
        return

    if not (type(produce_column) == int and type(price_column) == int):
        print("Column values should be integers only.")
        return

    if not type(produce_price) == dict:
        print("'produce_price' parameter should be a dictionary.")
        return


    # open Excel file
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    # make sure produce names in lowercase for further comparison
    produce_price = {k.lower() : v for k, v in produce_price.items()}

    # replace prices for required items
    for cell in list(sheet.columns)[produce_column-1]:
        produce = cell.value.lower()
        current_price = sheet.cell(row=cell.row, column=price_column).value

        if produce in produce_price.keys() and current_price != produce_price[produce]:
            new_price = produce_price[produce]
            sheet[get_column_letter(price_column)+str(cell.row)] = new_price
            print(f"Updated price for [{produce}] from {current_price} to {new_price}.")

    # save results in a new file
    wb.save(new_file_name)
    print(f"See results in '{new_file_name}'.")


file_for_update = "produce_sales.xlsx"
price_correction = {"GARLIC": 2, "CELERY": 3, "LEMONS": 4}
results_file_name = "produce_sales_updated.xlsx"

update_produce_price(file_for_update, 1, 2, price_correction, results_file_name)

