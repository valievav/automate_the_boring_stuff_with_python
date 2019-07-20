'''
Write a program that performs the tasks of the text_files_to_spreadsheet.py
in reverse order: The program should open a spreadsheet and write the cells of column A
into one text file, the cells of column B into another text file, and so on.
'''

import openpyxl, os


def spreadsheet_to_text_files(xlsx_file, txt_files):
    """
    Takes each column of the .xlsx file and records its data into separate .txt file.\n
    :param xlsx_file: .xlsx file
    :param txt_files: list with .txt files
    :return:
    """

    wb = openpyxl.load_workbook(xlsx_file)
    print(f"File for processing - '{xlsx_file}'")
    sheet = wb.active
    max_column = sheet.max_column
    max_row = sheet.max_row

    # iterate through each column and record data into separate txt file
    for col in range(1, max_column+1):
        file_name = txt_files[col-1]
        text_file = open(os.path.join(os.getcwd(), file_name), "w")
        column_data_obj = list(sheet.columns)[col-1]  # read all column data

        for row in range(1, max_row+1):
            if column_data_obj[row-1].value:  # skipping None rows
                text_file.write(column_data_obj[row-1].value+"\n")

        text_file.close()
        print(f"Populated file '{file_name}' with values from column {col}.")


main_file = "spreadsheet_to_text_files_file.xlsx"
result_files = ['spreadsheet_to_text_files_result1.txt', 'spreadsheet_to_text_files_result2.txt']

spreadsheet_to_text_files(main_file, result_files)

