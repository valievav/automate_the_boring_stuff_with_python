# This is support for the blank_row_inserter program.
# It populates values in the blank_row_file.xlsx file.

import openpyxl
from openpyxl.utils.cell import get_column_letter


def populate_values(file_name, max_row=10, max_column=10):
    """
    Populates cells in .xlsx file. If max row and columns are not specified, it uses default values.\n
    :param file_name: .xlsx extension
    :param max_row: int, optional
    :param max_column: int, optional
    :return:
    """

    wb = openpyxl.Workbook()
    sheet = wb.active

    # fill in all cells withing max_row & max_column area
    for row in range(1, max_row+1):
        for column in range(1, max_column+1):
            sheet[get_column_letter(column)+str(row)] = "packman"

    wb.save(file_name)


file = "blank_row_file.xlsx"
max_row_number = 15
max_column_number = 15

populate_values(file, max_row_number, max_column_number)

