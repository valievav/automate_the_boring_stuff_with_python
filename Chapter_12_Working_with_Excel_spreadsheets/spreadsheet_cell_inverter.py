'''
Write a program to invert the row and column of the cells in the spreadsheet.
For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa).
This should be done for all cells in the spreadsheet.
'''

import openpyxl
from openpyxl.utils.cell import get_column_letter


def invert_spreadsheet_cell(file_before, file_after=None):
    """
    Inverts data from the provided file into the same file or a new one.\n
    :param file_before: .xlsx extension
    :param file_after: .xlsx extension, optional
    :return:
    """

    # open spreadsheet
    wb = openpyxl.load_workbook(file_before)
    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column

    # extract data into a list and clean up the cells
    spreadsheet_data = []

    for row in range(1, max_row+1):  # 17
        row_list = []
        for col in range(1, max_col+1):  # 3
            value = sheet[get_column_letter(col)+str(row)].value
            row_list.append(value)
            sheet[get_column_letter(col) + str(row)] = ""  # clean cell after recording the value
        spreadsheet_data.append(row_list) # storing data from one row in one sublist

    # invert data and record into the spreadsheet
    for col in range(1, len(spreadsheet_data) + 1):  # 17
        for row in range(1, len(spreadsheet_data[0]) + 1):  # 3
            sheet[get_column_letter(col)+str(row)] = spreadsheet_data[col-1][row-1]

    # define whether to record data in a new file or the same
    if file_after:
        file = file_after
    else:
        file = file_before

    wb.save(file)
    print(f"Inverted data from '{file_before}'.\nFile with the result - '{file}'")


file_initial = "spreadsheet_file_inverter_file.xlsx"
file_result = "spreadsheet_file_inverter_file_after.xlsx"
invert_spreadsheet_cell(file_initial, file_result)

