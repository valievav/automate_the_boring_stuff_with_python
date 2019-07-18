'''
Write a program to read in the contents of several text files (you can make
the text files yourself) and insert those contents into a spreadsheet, with
one line of text per row. The lines of the first text file will be in the cells of
column A, the lines of the second text file will be in the cells of column B, and so on.
'''

import openpyxl
from openpyxl.utils.cell import get_column_letter


def text_files_to_spreadsheet(text_files):
    """
    Creates .xlsx file and populates each column with .txt file records.\n
    List of .txt files is passed as a param to the function.\n
    :param text_files: list with .txt file names
    :return: None
    """

    wb = openpyxl.Workbook()
    sheet = wb.active

    def recording_values():
        for item in text_files:
            file = open(item)
            lines = file.readlines()
            for line in lines:
                column = text_files.index(item)
                row = lines.index(line)
                sheet[get_column_letter(column+1)+str(row+1)] = line

    # alternative memory-efficient option
    def recording_values_memory_efficient():
        for col in range(len(text_files)):
            file = open(text_files[col])
            lines = file.readlines()
            for row in range(len(lines)):
                sheet[get_column_letter(col+1)+str(row+1)] = lines[row]

    recording_values()
    # recording_values_memory_efficient()

    wb.save("text_files_to_spreadsheet_result.xlsx")


text_files_scope = ['test_file_to_spreadsheet_file1.txt', 'test_file_to_spreadsheet_file2.txt']
text_files_to_spreadsheet(text_files_scope)


