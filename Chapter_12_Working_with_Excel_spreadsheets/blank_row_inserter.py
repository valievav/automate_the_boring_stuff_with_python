'''
Create a program blankRowInserter.py that takes two integers and a filename
string as command line arguments. Let’s call the first integer N and the second
integer M. Starting at row N, the program should insert M blank rows
into the spreadsheet.
Example of the program: py blank_row_inserter.py 3 2 myProduce.xlsx
'''

import sys, openpyxl
from openpyxl.utils.cell import get_column_letter
from blank_row_file_preparation import populate_values


def insert_blank_row(start_row, rows_to_insert, file_name):
    """
    Inserts empty rows in an Excel file.\n
    Parameters like start_row, rows_to_insert, file_name are received from the command line.\n
    :param start_row: int
    :param rows_to_insert: int
    :param file_name: .xlsx extension
    :return:
    """

    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    max_col = sheet.max_column

    # inserting empty values
    for row in range(start_row, start_row+rows_to_insert):
        for col in range(1, max_col+1):
            sheet[get_column_letter(col)+str(row)] = ""

    print(f"Inserted {rows_to_insert} blank rows starting from row {start_row} in '{file_name}'")
    wb.save(file_name)


# use either system args or default ones
try:
    sys.argv[1]
except IndexError:
    start_row_number = 5
    blank_rows_number = 2
    file = "blank_row_file.xlsx"
else:
    start_row_number = int(sys.argv[1])
    blank_rows_number = int(sys.argv[2])
    file = sys.argv[3]


populate_values(file)
insert_blank_row(start_row_number, blank_rows_number, file)

