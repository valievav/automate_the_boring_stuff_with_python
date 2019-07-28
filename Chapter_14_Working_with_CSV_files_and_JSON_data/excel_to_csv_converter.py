'''
Write a program that reads all the Excel files in the current working directory and outputs them as CSV files.
A single Excel file might contain multiple sheets; you’ll have to create one CSV file per sheet.
The filenames of the CSV files should be <excel filename>_<sheet title>.csv.
'''

import openpyxl, csv, os


def convert_excel_to_csv(cwd=None):
    """
    Converts .xlsx file sheets into separate .csv files.\n
    Csv files have name in format <excel filename>_<sheet title>.csv
    :param cwd: absolute path, optional param
    :return:
    """

    # switch to cwd if passed as param
    if cwd:
        os.chdir(cwd)

    # search for xlsx file in cmd
    for file in os.listdir('.'):
        if file.endswith(".xlsx"):

            wb = openpyxl.load_workbook(file)
            sheets = wb.sheetnames

            # iterate through file's sheets
            for sheet in sheets:
                active_sheet = wb[sheet]

                # skip empty sheet
                if active_sheet.max_row == 1 and active_sheet.max_column ==1 and not active_sheet["A1"].value:
                    continue
                else:
                    # create csv file
                    csv_file_name_parts = file.title().split('.')  # splitting file name to extract part w/o extension
                    csv_file_name = f"{csv_file_name_parts[0].lower()}_{sheet.title().lower()}.csv"
                    csv_file = open(csv_file_name, 'w', newline='')
                    writer = csv.writer(csv_file)

                    # record values from excel file
                    writer.writerows(active_sheet.values)
                    csv_file.close()
                    print(f"Created file '{csv_file_name}'")


working_dir = "D:\\Practice Python\\Chapter 14"
convert_excel_to_csv()

